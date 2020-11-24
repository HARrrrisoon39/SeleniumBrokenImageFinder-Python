import requests
from selenium import webdriver
import openpyxl
wb = openpyxl.Workbook()  

sheet = wb.active 
options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_argument('disable-infobars')
driver = webdriver.Chrome(chrome_options=options,executable_path=r'C:\utility\chromedriver.exe')
driver.get('https://cbdbene.com/')
images = driver.find_elements_by_tag_name('img')
for image in images:
    if image.get_attribute('naturalWidth') == "0":
        print(image.get_attribute('src'))
        c1 = sheet.cell(row = 1, column = 1) 
        c1.value = "1"
        c2 = sheet.cell(row= 1 , column = 2) 
        c2.value = image.get_attribute('src') 
        wb.save('xlwt example.xls')